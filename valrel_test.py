import os
import pandas as pd
import openpyxl
from pydbml import PyDBML

def obj_master_parent_parse(dbml_file: str, excel_file: str, object_api_name: str):
    try:
        # Check if the Excel file exists
        if not os.path.exists(excel_file):
            print(f"Error: File {excel_file} not found.")
            return

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)

        # Check if the object sheet exists
        if object_api_name not in workbook.sheetnames:
            print(f"Error: Sheet {object_api_name} not found in {excel_file}.")
            return

        # Load the object sheet
        sheet = workbook[object_api_name]
        data = pd.DataFrame(sheet.values)
        headers = data.iloc[0].tolist()
        data = data[1:]
        data.columns = headers

        # Hardcoded DBML example for testing
        table_name = "OrderItem"
        hardcoded_relationships = [
            ("OrderId", "Order", "Id"),
            ("PricebookEntryId", "PricebookEntry", "Id")
        ]

        # Simulating table relationships
        relationships = hardcoded_relationships if object_api_name == table_name else []

        if not relationships:
            print(f"No relationships defined for {object_api_name}.")
            return

        unmatched = {}

        for column_name, ref_table, ref_column in relationships:
            if column_name not in headers:
                print(f"Missing required relationship: {object_api_name}.{column_name}")
                continue

            # Fetch referenced table
            if ref_table not in workbook.sheetnames:
                print(f"Referenced table {ref_table} not found in workbook.")
                unmatched[column_name] = "Table not found"
                continue

            ref_sheet = workbook[ref_table]
            ref_data = pd.DataFrame(ref_sheet.values)
            ref_headers = ref_data.iloc[0].tolist()
            ref_data = ref_data[1:]
            ref_data.columns = ref_headers

            if ref_column not in ref_headers:
                print(f"Column {ref_column} not found in referenced table {ref_table}.")
                unmatched[column_name] = "Column not found"
                continue

            # Check for matches
            ref_ids = ref_data[ref_column].dropna().unique()
            sheet_ids = data[column_name].dropna().unique()
            matched_ids = set(sheet_ids).intersection(ref_ids)
            unmatched_ids = set(sheet_ids).difference(ref_ids)

            print(f"On Sheet {object_api_name}.{column_name} has {len(sheet_ids)} Id values.")
            print(f"In Sheet {ref_table}.{ref_column} finds {len(matched_ids)} of {len(sheet_ids)} matching Id's.")

            if unmatched_ids:
                unmatched[column_name] = unmatched_ids

        if unmatched:
            print("\nValidation failed:")
            for column, issue in unmatched.items():
                if isinstance(issue, set):
                    rows = [i + 2 for i, val in enumerate(data[column_name]) if val in issue]
                    print(f"  {object_api_name}.{column} has unmatched Ids at rows: {rows}.")
                else:
                    print(f"  {column}: {issue}")
            print("Object Parent Data relationship validation IS NOT met!")
        else:
            print("Parent Data relationship validated successfully.")
    except Exception as e:
        print(f"Critical Error: {e}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Validate Object Parent Relationships")
    parser.add_argument("-t", "--target", required=True, help="Path to the Excel file")
    parser.add_argument("-o", "--object", required=True, help="Object API Name")
    parser.add_argument("-d", "--dbml", required=False, help="Path to the DBML file")

    args = parser.parse_args()
    obj_master_parent_parse(args.dbml, args.target, args.object)