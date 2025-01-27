import os
import pandas as pd
import openpyxl
from pydbml import PyDBML

def obj_master_parent_parse(dbml_file: str, excel_file: str, object_api_name: str):
    try:
        print(f"Debug: Starting validation for {object_api_name} using {dbml_file} and {excel_file}")

        # Check if the Excel file exists
        if not os.path.exists(excel_file):
            print(f"Error: File {excel_file} not found.")
            return

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        print(f"Debug: Workbook {excel_file} loaded successfully with sheets: {workbook.sheetnames}")

        # Check if the object sheet exists
        if object_api_name not in workbook.sheetnames:
            print(f"Error: Sheet {object_api_name} not found in {excel_file}.")
            return

        # Load the object sheet
        sheet = workbook[object_api_name]
        data = pd.DataFrame(sheet.values)
        headers = data.iloc[0].tolist()
        data = data[1:]
        data.columns = headers
        print(f"Debug: Sheet {object_api_name} loaded successfully with headers: {headers}")
        print(f"Debug: Data in sheet:\n{data.head()}")

        # Parse the DBML file
        with open(dbml_file, "r") as file:
            dbml_content = PyDBML(file.read())
        print(f"Debug: DBML file {dbml_file} parsed successfully.")

        # Extract the table for the given object
        table = next((tbl for tbl in dbml_content.tables if tbl.name == object_api_name), None)
        if not table:
            print(f"Error: No table found for {object_api_name} in {dbml_file}. Available tables: {[tbl.name for tbl in dbml_content.tables]}")
            return
        print(f"Debug: Table {object_api_name} found in DBML file.")

        # Parse the DBML file with property support
        with open(dbml_file, "r") as file:
            dbml_content = PyDBML(file.read(), allow_properties=True)
        print(f"Debug: DBML file {dbml_file} parsed successfully with properties enabled.")


        # Extract relationships from the table using properties
        relationships = []
        for column in table.columns:
            if 'ref' in column.properties and column.not_null:
                # Extract referenced table and column from the `ref` property
                ref_table, ref_column = column.properties['ref'].split('.')
                relationships.append((column.name, ref_table.strip(), ref_column.strip()))

        print(f"Debug: Relationships extracted for {object_api_name}: {relationships}")

        if not relationships:
            print(f"No required relationships with [ref: > ... AND not null] found for {object_api_name}.")
            return

        unmatched = {}

        for column_name, ref_table, ref_column in relationships:
            print(f"Debug: Validating relationship {object_api_name}.{column_name} -> {ref_table}.{ref_column}")

            if column_name not in headers:
                print(f"Error: Required relationship {object_api_name}.{column_name} not found in sheet.")
                continue

            # Fetch referenced table
            if ref_table not in workbook.sheetnames:
                print(f"Error: Referenced table {ref_table} not found in workbook.")
                unmatched[column_name] = "Table not found"
                continue

            ref_sheet = workbook[ref_table]
            ref_data = pd.DataFrame(ref_sheet.values)
            ref_headers = ref_data.iloc[0].tolist()
            ref_data = ref_data[1:]
            ref_data.columns = ref_headers
            print(f"Debug: Referenced table {ref_table} loaded successfully with headers: {ref_headers}")
            print(f"Debug: Data in referenced table:\n{ref_data.head()}")

            if ref_column not in ref_headers:
                print(f"Error: Column {ref_column} not found in referenced table {ref_table}.")
                unmatched[column_name] = "Column not found"
                continue

            # Check for matches
            ref_ids = ref_data[ref_column].dropna().unique()
            sheet_ids = data[column_name].dropna().unique()
            matched_ids = set(sheet_ids).intersection(ref_ids)
            unmatched_ids = set(sheet_ids).difference(ref_ids)

            print(f"On Sheet {object_api_name}.{column_name} has {len(sheet_ids)} Id values.")
            print(f"In Sheet {ref_table}.{ref_column} finds {len(matched_ids)} of {len(sheet_ids)} matching Id's.")

            if unmatched_ids:
                unmatched[column_name] = unmatched_ids

        if unmatched:
            print("\nValidation failed:")
            for column, issue in unmatched.items():
                if isinstance(issue, set):
                    rows = [i + 2 for i, val in enumerate(data[column_name]) if val in issue]
                    print(f"  {object_api_name}.{column} has unmatched Ids at rows: {rows}.")
                else:
                    print(f"  {column}: {issue}")
            print("Object Parent Data relationship validation IS NOT met!")
        else:
            print("Parent Data relationship validated based on .dbml")
    except Exception as e:
        print(f"Critical Error: {e}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Validate Object Parent Relationships")
    parser.add_argument("-t", "--target", required=True, help="Path to the Excel file")
    parser.add_argument("-o", "--object", required=True, help="Object API Name")
    parser.add_argument("-d", "--dbml", required=True, help="Path to the DBML file")

    args = parser.parse_args()

    print(f"Debug: Running validation with args {args}")
    obj_master_parent_parse(args.dbml, args.target, args.object)
    print("Debug: Validation completed.")
