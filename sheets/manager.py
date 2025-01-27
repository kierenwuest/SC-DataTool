from openpyxl import Workbook
from datetime import datetime

def flatten_record(record):
    """
    Flattens nested dictionary-like structures into a single-level dictionary.

    Args:
        record (dict): A dictionary that may contain nested dictionaries.

    Returns:
        dict: A flattened dictionary with all values converted to strings.
    """
    flattened = {}
    for key, value in record.items():
        if isinstance(value, dict):  # Handle nested dictionaries
            for sub_key, sub_value in value.items():
                flattened[f"{key}.{sub_key}"] = str(sub_value)  # Convert to string
        elif isinstance(value, (list, tuple)):  # Handle lists or tuples
            flattened[key] = ", ".join(map(str, value))  # Convert to a comma-separated string
        else:  # Handle scalar values
            flattened[key] = str(value) if value is not None else ""  # Convert to string
    return flattened

def create_workbook(data, filename, log_data):
    """
    Creates an Excel workbook with logs and data.

    Args:
        data (dict): Data for each table (table name as key, records as values).
        filename (str): Path to save the workbook.
        log_data (list of lists): Logs to include in the workbook's Summary tab.
    """
    workbook = Workbook()

    # Create the summary sheet for logs
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    # Write logs starting at row 1
    for row_idx, log_row in enumerate(log_data, start=1):
        for col_idx, value in enumerate(log_row, start=1):
            summary_sheet.cell(row=row_idx, column=col_idx, value=value)

    # Create sheets for each table's data
    for sheet_name, records in data.items():
        # Create a new sheet for the table
        sheet = workbook.create_sheet(title=sheet_name)

        if records:
            # Flatten records and extract headers, excluding 'attributes' columns
            flattened_records = [flatten_record(record) for record in records]
            headers = [key for key in flattened_records[0].keys() if key not in ['attributes.type', 'attributes.url'] and not key.endswith(".attributes")]
            sheet.append(headers)

            # Append flattened records
            for record in flattened_records:
                row = [record.get(header, "") for header in headers]
                sheet.append(row)

    # Save the workbook
    workbook.save(filename)
